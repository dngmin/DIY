from openpyxl import load_workbook

class Para:
    def __init__(self,file):
        self.file = file

    def nose(self):
        wb = load_workbook(self.file)
        ws = wb.active

        nose_para = {
            'shape_nose'     : ws['D6'].value,
            'material'       : ws['D7'].value,
            'shape_coef'     : ws['D8'].value,
            'length_nose'    : ws['D9'].value  / 10,
            'diameter_nose'  : ws['D10'].value  / 10,
            'thickness_nose' : ws['D11'].value /10
        }
        return nose_para
    
    def body(self):
        wb = load_workbook(self.file)
        ws = wb.active

        body_para = {
            'quantity_body' : ws['D24'].value,
            'diameter_body' : ws['D25'].value / 10
        }
        for i in range(1,body_para['quantity_body']+1):
            bodytube = {
                'lenght_body' : ws[f'D{27+(i-1)*4}'].value / 10,
                'material'    : ws[f'D{28+(i-1)*4}'].value,
                'inner_diameter' : ws[f'D{29+(i-1)*4}'].value / 10
            }
            body_para[f'body{i}'] = bodytube
        return body_para

    def fin(self):
        wb = load_workbook(self.file)
        ws = wb.active

        fin_para = {
            'shape_fin'          : ws['D13'].value,
            'quantity_fin'       : ws['D14'].value,
            'inclination'        : ws['D15'].value,
            'root_chord'         : ws['D16'].value / 10,
            'tip_chord'          : ws['D17'].value / 10,
            'span'               : ws['D18'].value / 10,
            'leading_edge_chord' : ws['D19'].value / 10,
            'offset_fin'         : ws['D20'].value / -10,
            'thickness_fin'      : ws['D21'].value / 10,
            'material'           : ws['D22'].value
        }
        return fin_para